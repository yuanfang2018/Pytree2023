import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import os

# 爬取网页
url = 'https://www.exfeed.jp/top.php'
response = requests.get(url)
soup = BeautifulSoup(response.content, 'html.parser')

# 创建Excel文件
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = 'Exfeed'

# 写入表头
sheet['A1'] = '作品名'
sheet['B1'] = '作品封面'
sheet['C1'] = '发售日'
sheet['D1'] = '番号'

# 设置表头格式
for col in sheet.columns:
    column_letter = get_column_letter(col[0].column)
    sheet.column_dimensions[column_letter].width = 20
    for cell in col:
        cell.alignment = Alignment(horizontal='center', vertical='center')

# 获取作品信息
works = soup.find_all('div', class_='work')

# 写入作品信息
for i, work in enumerate(works):
    name = work.find('div', class_='work_name').text.strip()
    cover_url = work.find('img')['src']
    release_date = work.find('div', class_='work_date').text.strip()
    code = work.find('div', class_='work_code').text.strip()
    sheet.cell(row=i+2, column=1, value=name)
    sheet.cell(row=i+2, column=2).value = cover_url
    sheet.cell(row=i+2, column=3, value=release_date)
    sheet.cell(row=i+2, column=4, value=code)

# 保存Excel文件
filename = 'Exfeed.xlsx'
wb.save(filename)
print(f'{os.path.abspath(filename)}已保存')
